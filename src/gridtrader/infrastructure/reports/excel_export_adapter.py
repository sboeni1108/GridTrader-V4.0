"""
Adapter zwischen Widget Dictionary-Results und Excel-Export
Konvertiert Dictionary-Daten in Excel ohne Domain-Objekte
"""
from typing import Dict, Optional
from pathlib import Path
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExportAdapter:
    """Adaptiert Dictionary-Results für Excel-Export"""

    def export_backtest_results(
        self,
        results: Dict[str, Dict],
        scenarios: Dict[str, Dict],
        symbol: str,
        timeframe_days: int,
        candle_minutes: int,
        historical_data: Optional[pd.DataFrame] = None,
        output_dir: str = "reports/backtest"
    ) -> Path:
        """
        Exportiert Multi-Szenario Backtest als Excel
        Konvertiert Dictionary-Daten in Excel ohne Domain-Objekte

        Args:
            results: Dictionary mit Backtest-Resultaten {scenario_name: result_dict}
            scenarios: Dictionary mit Szenario-Konfigurationen
            symbol: Handelssymbol
            timeframe_days: Analysezeitraum in Tagen
            candle_minutes: Kerzen-Timeframe in Minuten
            historical_data: Optional historische Daten
            output_dir: Ausgabeverzeichnis

        Returns:
            Pfad zur erstellten Excel-Datei
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = output_path / f"Backtest_{symbol}_{timestamp}.xlsx"

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Übersicht
            self._write_overview(writer, results, scenarios, symbol, timeframe_days, candle_minutes, historical_data)

            # 2. TOP 10
            self._write_top10(writer, results, scenarios)

            # 3. Alle Resultate
            self._write_all_results(writer, results, scenarios)

            # 4. Details LONG
            self._write_details_long(writer, results, scenarios)

            # 5. Details SHORT
            self._write_details_short(writer, results, scenarios)

            # 6. Gegenüberstellung
            self._write_comparison(writer, results, scenarios)

            # 7. Formatierung anwenden
            self._apply_formatting(writer.book)

        print(f"📊 Excel-Report erstellt: {filename}")
        return filename

    def _write_overview(self, writer, results, scenarios, symbol, timeframe_days, candle_minutes, historical_data):
        """Schreibt Übersichtsblatt mit vollständigen Metriken wie im Original"""

        # Beste Konfiguration finden
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1].get('net_pnl', 0),
            reverse=True
        )
        best_name, best_result = sorted_results[0] if sorted_results else (None, {})
        best_scenario = scenarios.get(best_name, {}) if best_name else {}

        # Kerzen-Statistiken berechnen
        kerzen_stats = self._calculate_candle_stats(historical_data) if historical_data is not None else {}

        # Daten für Übersicht - EXAKT wie im Original
        data = [
            ['Parameter', 'Wert'],
            ['GRID TRADING OPTIMIERUNG - ERGEBNISSE', ''],
            ['', ''],
            ['📊 ANALYSE-UMFANG', ''],
            ['Symbol', symbol],
            ['Analysezeitraum', f'{timeframe_days} Tage'],
            ['Timeframe', f'{candle_minutes} Minuten'],
            ['Konfigurationen getestet', len(results)],
            ['', ''],

            ['📈 KERZEN-STATISTIKEN', ''],
            ['Ø Tages-Range', f"{kerzen_stats.get('mean_daily_range', 0):.2f}%"],
            ['Ø Kerzen-Range', f"{kerzen_stats.get('mean_candle_range', 0):.2f}%"],
            ['50% Perzentil', f"{kerzen_stats.get('percentile_50', 0):.2f}%"],
            ['75% Perzentil', f"{kerzen_stats.get('percentile_75', 0):.2f}%"],
            ['Typischer Rebound', f"{kerzen_stats.get('typical_rebound', 0):.2f}%"],
            ['Bullish Kerzen', f"{kerzen_stats.get('bullish_percentage', 0):.1f}%"],
            ['Extreme Tage', f"{kerzen_stats.get('extreme_days_pct', 0):.1f}%"],
            ['', ''],

            ['🏆 BESTE KONFIGURATION', ''],
            ['Konfigurations-ID', best_name or 'N/A'],
            ['Basis-Typ', best_scenario.get('type', 'N/A')],
            ['Sharpe Ratio', f"{best_result.get('sharpe_ratio', 0):.3f}"],
            ['Win Rate', f"{best_result.get('win_rate', 0):.1f}%"],
            ['ROI', f"{best_result.get('roi', 0):.2f}%"],
            ['', ''],

            ['💰 PERFORMANCE (Beste Konfiguration)', ''],
            ['Erfolgreiche Trades', best_result.get('trades', 0)],
            ['Brutto-Gewinn', f"${best_result.get('pnl_usd', 0):,.2f}"],
            ['Gebühren Total', f"${best_result.get('commission', 0):,.2f}"],
            ['Netto-Gewinn', f"${best_result.get('net_pnl', 0):,.2f}"],
            ['Ø Tagesgewinn', f"${best_result.get('net_pnl', 0) / max(timeframe_days, 1):.2f}"],
            ['Max Drawdown', f"{best_result.get('max_drawdown', 0) * 100:.2f}%"],
            ['Kapital-Effizienz', f"{best_result.get('capital_efficiency', 0):.2f}%"],
            ['', ''],

            ['📦 OFFENE POSITIONEN (Ende)', ''],
            ['Anzahl Aktien', best_result.get('remaining_shares', 0)],
            ['Investiertes Kapital', f"${best_result.get('avg_entry_price', 0) * best_result.get('remaining_shares', 0):,.2f}"],
            ['Unrealisierter P&L', f"${best_result.get('unrealized_pnl', 0):,.2f}"],
            ['', ''],

            ['⚠️ RISIKO-METRIKEN', ''],
            ['Notfall-Stops ausgelöst', best_result.get('emergency_stops', 0)],
            ['Ø Drawdown', f"{best_result.get('avg_drawdown', 0):.2f}%"],
            ['Volatilität (Std)', f"${best_result.get('volatility', 0):.2f}"],
            ['', ''],

            ['📅 ERSTELLT', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]

        # DataFrame erstellen und schreiben
        df = pd.DataFrame(data[1:], columns=data[0])
        df.to_excel(writer, sheet_name='Übersicht', index=False)

    def _calculate_candle_stats(self, historical_data):
        """Berechnet erweiterte Kerzen-Statistiken"""
        if historical_data is None or historical_data.empty:
            return {}

        df = historical_data

        # Tages-Statistiken
        daily_stats = df.groupby(df.index.date).agg({
            'open': 'first',
            'high': 'max',
            'low': 'min',
            'close': 'last'
        })

        daily_stats['range'] = (daily_stats['high'] - daily_stats['low']) / daily_stats['open'] * 100

        # Kerzen-Statistiken
        df['candle_range'] = (df['high'] - df['low']) / df['open'] * 100
        df['bullish'] = df['close'] > df['open']

        return {
            'mean_daily_range': daily_stats['range'].mean(),
            'mean_candle_range': df['candle_range'].mean(),
            'percentile_50': daily_stats['range'].quantile(0.5),
            'percentile_75': daily_stats['range'].quantile(0.75),
            'typical_rebound': daily_stats['range'].quantile(0.6),
            'bullish_percentage': df['bullish'].mean() * 100,
            'extreme_days_pct': (daily_stats['range'] > daily_stats['range'].quantile(0.9)).mean() * 100
        }

    def _write_top10(self, writer, results, scenarios):
        """Schreibt TOP 10 Blatt"""
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1].get('net_pnl', 0),
            reverse=True
        )

        top10_data = []
        for rank, (name, result) in enumerate(sorted_results[:10], 1):
            config = scenarios.get(name, {})
            top10_data.append({
                'Rang': rank,
                'Szenario': name,
                'Typ': config.get('type', 'N/A'),
                'Initial Capital': f"${result.get('initial_capital', 0):,.2f}",
                'Net P&L %': f"{result.get('pnl_percent', 0):.2f}%",
                'Net P&L $': f"${result.get('net_pnl', 0):,.2f}",
                'Realized P&L $': f"${result.get('pnl_usd', 0):,.2f}",
                'Win Rate %': f"{result.get('win_rate', 0):.1f}%",
                'Max DD %': f"{result.get('max_drawdown', 0):.2f}%",
                'Trades': result.get('trades', 0),
                'Rest Aktien': result.get('remaining_shares', 0),
                'Kommission $': f"${result.get('commission', 0):,.2f}"
            })

        pd.DataFrame(top10_data).to_excel(writer, sheet_name='TOP 10', index=False)

    def _write_all_results(self, writer, results, scenarios):
        """Schreibt alle Resultate mit Config-Parametern"""
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1].get('net_pnl', 0),
            reverse=True
        )

        all_data = []
        for name, result in sorted_results:
            config = scenarios.get(name, {})
            all_data.append({
                'Szenario': name,
                'Typ': config.get('type', 'N/A'),
                'Shares': config.get('shares', 0),
                'Step %': config.get('step', 0),
                'Exit %': config.get('exit', 0),
                'Levels': config.get('levels', 0),
                'Initial Capital': f"${result.get('initial_capital', 0):,.2f}",
                'Net P&L %': f"{result.get('pnl_percent', 0):.2f}%",
                'Net P&L $': f"${result.get('net_pnl', 0):,.2f}",
                'Realized P&L $': f"${result.get('pnl_usd', 0):,.2f}",
                'Unrealized P&L $': f"${result.get('unrealized_pnl', 0):,.2f}",
                'Win Rate %': f"{result.get('win_rate', 0):.1f}%",
                'Max DD %': f"{result.get('max_drawdown', 0):.2f}%",
                'Trades': result.get('trades', 0),
                'Rest Aktien': result.get('remaining_shares', 0),
                'Kommission $': f"${result.get('commission', 0):,.2f}"
            })

        pd.DataFrame(all_data).to_excel(writer, sheet_name='Alle Resultate', index=False)

    def _write_details_long(self, writer, results, scenarios):
        """Schreibt Details für LONG-Szenarien"""
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1].get('net_pnl', 0),
            reverse=True
        )

        long_data = []
        for name, result in sorted_results:
            config = scenarios.get(name, {})
            if config.get('type') == 'LONG':
                long_data.append({
                    'Szenario': name,
                    'Shares': config.get('shares', 0),
                    'Step %': config.get('step', 0),
                    'Exit %': config.get('exit', 0),
                    'Levels': config.get('levels', 0),
                    'Initial Capital': f"${result.get('initial_capital', 0):,.2f}",
                    'Net P&L %': f"{result.get('pnl_percent', 0):.2f}%",
                    'Net P&L $': f"${result.get('net_pnl', 0):,.2f}",
                    'Realized P&L $': f"${result.get('pnl_usd', 0):,.2f}",
                    'Unrealized P&L $': f"${result.get('unrealized_pnl', 0):,.2f}",
                    'Win Rate %': f"{result.get('win_rate', 0):.1f}%",
                    'Max DD %': f"{result.get('max_drawdown', 0):.2f}%",
                    'Trades': result.get('trades', 0),
                    'Rest Aktien': result.get('remaining_shares', 0)
                })

        if long_data:
            pd.DataFrame(long_data).to_excel(writer, sheet_name='Details LONG', index=False)

    def _write_details_short(self, writer, results, scenarios):
        """Schreibt Details für SHORT-Szenarien"""
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1].get('net_pnl', 0),
            reverse=True
        )

        short_data = []
        for name, result in sorted_results:
            config = scenarios.get(name, {})
            if config.get('type') == 'SHORT':
                short_data.append({
                    'Szenario': name,
                    'Shares': config.get('shares', 0),
                    'Step %': config.get('step', 0),
                    'Exit %': config.get('exit', 0),
                    'Levels': config.get('levels', 0),
                    'Initial Capital': f"${result.get('initial_capital', 0):,.2f}",
                    'Net P&L %': f"{result.get('pnl_percent', 0):.2f}%",
                    'Net P&L $': f"${result.get('net_pnl', 0):,.2f}",
                    'Realized P&L $': f"${result.get('pnl_usd', 0):,.2f}",
                    'Unrealized P&L $': f"${result.get('unrealized_pnl', 0):,.2f}",
                    'Win Rate %': f"{result.get('win_rate', 0):.1f}%",
                    'Max DD %': f"{result.get('max_drawdown', 0):.2f}%",
                    'Trades': result.get('trades', 0),
                    'Rest Aktien': result.get('remaining_shares', 0)
                })

        if short_data:
            pd.DataFrame(short_data).to_excel(writer, sheet_name='Details SHORT', index=False)

    def _write_comparison(self, writer, results, scenarios):
        """Schreibt Gegenüberstellung LONG vs SHORT"""
        long_results = {
            name: result
            for name, result in results.items()
            if scenarios.get(name, {}).get('type') == 'LONG'
        }
        short_results = {
            name: result
            for name, result in results.items()
            if scenarios.get(name, {}).get('type') == 'SHORT'
        }

        best_long = max(long_results.items(), key=lambda x: x[1].get('net_pnl', 0)) if long_results else (None, {})
        best_short = max(short_results.items(), key=lambda x: x[1].get('net_pnl', 0)) if short_results else (None, {})

        comparison_data = [
            ['Metrik', 'LONG (Beste)', 'SHORT (Beste)'],
            ['Szenario', best_long[0] or 'N/A', best_short[0] or 'N/A'],
            ['Net P&L $', f"${best_long[1].get('net_pnl', 0):,.2f}", f"${best_short[1].get('net_pnl', 0):,.2f}"],
            ['Net P&L %', f"{best_long[1].get('pnl_percent', 0):.2f}%", f"{best_short[1].get('pnl_percent', 0):.2f}%"],
            ['Win Rate %', f"{best_long[1].get('win_rate', 0):.1f}%", f"{best_short[1].get('win_rate', 0):.1f}%"],
            ['Max DD %', f"{best_long[1].get('max_drawdown', 0):.2f}%", f"{best_short[1].get('max_drawdown', 0):.2f}%"],
            ['Trades', best_long[1].get('trades', 0), best_short[1].get('trades', 0)],
            ['Rest Aktien', best_long[1].get('remaining_shares', 0), best_short[1].get('remaining_shares', 0)],
            ['Kommission $', f"${best_long[1].get('commission', 0):,.2f}", f"${best_short[1].get('commission', 0):,.2f}"]
        ]

        pd.DataFrame(comparison_data).to_excel(writer, sheet_name='Gegenüberstellung', index=False, header=False)

    def _apply_formatting(self, wb):
        """Wendet professionelle Formatierung auf alle Sheets an"""

        # Definiere Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        subheader_font = Font(bold=True, size=10)
        subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        title_font = Font(bold=True, size=14, color="1F4E78")

        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Positive/Negative Farben
        positive_font = Font(color="006100")
        negative_font = Font(color="9C0006")

        # Top 3 Hervorhebung
        top_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        # Formatiere jedes Sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # === SPEZIELLE FORMATIERUNG PRO SHEET ===

            if sheet_name == 'Übersicht':
                # Titel-Zeile formatieren
                if sheet['A1'].value:
                    sheet['A1'].font = Font(bold=True, size=12)

                # Haupttitel
                if sheet.max_row >= 2 and sheet['A2'].value:
                    sheet['A2'].font = title_font

                # Kategorien hervorheben (Zeilen mit Emojis)
                for row_idx in range(1, sheet.max_row + 1):
                    cell_value = str(sheet.cell(row=row_idx, column=1).value or '')

                    # Icons am Anfang = Kategorie
                    if cell_value.startswith(('📊', '📈', '🏆', '💰', '📦', '⚠️', '📅')):
                        sheet.cell(row=row_idx, column=1).font = subheader_font
                        sheet.cell(row=row_idx, column=1).fill = subheader_fill
                        sheet.cell(row=row_idx, column=2).fill = subheader_fill

            elif sheet_name in ['TOP 10', 'Alle Resultate', 'Details LONG', 'Details SHORT']:
                # Header formatieren (erste Zeile)
                for cell in sheet[1]:
                    if cell.value:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                # Top 3 Einträge hervorheben (nur bei TOP 10)
                if sheet_name == 'TOP 10':
                    for row_idx in range(2, min(5, sheet.max_row + 1)):
                        for cell in sheet[row_idx]:
                            cell.fill = top_fill

                # Freeze erste Zeile
                sheet.freeze_panes = 'A2'

                # Conditional Formatting für P&L Spalten
                for row_idx in range(2, sheet.max_row + 1):
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            value_str = str(cell.value)
                            # Prüfe auf negative Werte
                            if value_str.startswith('-$') or value_str.startswith('-') and '$' in value_str:
                                cell.font = negative_font
                            elif '$' in value_str and not value_str.startswith('-'):
                                # Prüfe ob es ein negativer Prozent ist
                                pass
                            # Prüfe Prozent-Werte
                            if '%' in value_str:
                                try:
                                    num = float(value_str.replace('%', '').replace(',', '.'))
                                    if num < 0:
                                        cell.font = negative_font
                                except:
                                    pass

            elif sheet_name == 'Gegenüberstellung':
                # Erste Zeile als Header
                for cell in sheet[1]:
                    if cell.value:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                # Erste Spalte als Labels
                for row_idx in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=1)
                    if cell.value:
                        cell.font = Font(bold=True)

            # === ALLGEMEINE FORMATIERUNG ===

            # Auto-Spaltenbreite
            for column in sheet.columns:
                max_length = 0

                # Hole column_letter sicher
                first_cell = column[0]

                if hasattr(first_cell, 'column_letter'):
                    column_letter = first_cell.column_letter
                elif hasattr(first_cell, 'column'):
                    column_letter = get_column_letter(first_cell.column)
                else:
                    continue

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                            # Rahmen hinzufügen (nicht für Übersicht-Titel)
                            if sheet_name != 'Übersicht' or cell.row > 3:
                                cell.border = border
                    except:
                        pass

                # Setze Spaltenbreite (min 10, max 50)
                if max_length > 0:
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    try:
                        sheet.column_dimensions[column_letter].width = adjusted_width
                    except:
                        pass

            # Zeilenhöhe für Header
            if sheet.max_row > 0:
                sheet.row_dimensions[1].height = 25
