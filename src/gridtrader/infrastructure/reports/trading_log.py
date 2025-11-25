"""
Trading Log - Excel Export für GridTrader V3.0
Erstellt professionelle Excel-Logs für Trades und P/L im Schweizer Format
"""
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import locale


class TradingLogExporter:
    """
    Exportiert Trading-Daten in vier Excel-Dateien:
    1. Tägliche Trades (neue Datei bei jedem App-Start)
    2. Jährliche Trades (fortlaufend pro Jahr)
    3. Täglicher P/L (neue Datei bei jedem App-Start)
    4. Jährlicher P/L (fortlaufend pro Jahr)
    """

    def __init__(self, logs_dir: Path):
        """
        Args:
            logs_dir: Verzeichnis für Log-Dateien (~/.gridtrader/logs/)
        """
        self.logs_dir = Path(logs_dir)
        self.logs_dir.mkdir(parents=True, exist_ok=True)

        # Session-Start für Dateinamen
        self.session_start = datetime.now()
        self.session_date = self.session_start.strftime("%Y-%m-%d")
        self.session_time = self.session_start.strftime("%H-%M-%S")
        self.year = self.session_start.year

        # Dateinamen
        self.daily_trades_file = self.logs_dir / f"{self.session_date} {self.session_time} Tagestrades.xlsx"
        self.yearly_trades_file = self.logs_dir / f"{self.year} Jahrestrades.xlsx"
        self.daily_pl_file = self.logs_dir / f"{self.session_date} {self.session_time} Tages PL.xlsx"
        self.yearly_pl_file = self.logs_dir / f"{self.year} Jahres PL.xlsx"

        # Styles
        self._init_styles()

        # Versuche Schweizer Locale zu setzen
        try:
            locale.setlocale(locale.LC_ALL, 'de_CH.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'German_Switzerland.1252')
            except:
                pass

        # Initialisiere tägliche Dateien (bei jedem Start neu)
        self._init_daily_files()

        # Initialisiere jährliche Dateien (falls nicht vorhanden)
        self._init_yearly_files()

        # Tracking für P/L-Aggregation pro Symbol
        self.daily_symbol_stats: Dict[str, Dict] = {}

    def _init_styles(self):
        """Initialisiere Excel-Styles"""
        # Header Style
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Totals Style
        self.totals_font = Font(bold=True, size=11)
        self.totals_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Border
        self.border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Positive/Negative
        self.positive_font = Font(color="006100")
        self.negative_font = Font(color="9C0006")

    def _init_daily_files(self):
        """Erstelle neue tägliche Excel-Dateien"""
        # Tagestrades
        wb_trades = Workbook()
        ws_trades = wb_trades.active
        ws_trades.title = "Tagestrades"

        trade_headers = ["Datum/Zeit", "Symbol", "Side", "Anzahl", "Preis", "Kommission", "Total Kosten"]
        self._write_headers(ws_trades, trade_headers)
        wb_trades.save(self.daily_trades_file)

        # Tages P/L
        wb_pl = Workbook()
        ws_pl = wb_pl.active
        ws_pl.title = "Tages PL"

        pl_headers = ["Datum", "Symbol", "Total Aktien", "Total Trades", "Realisierter P/L", "Kommissionen", "Netto P/L"]
        self._write_headers(ws_pl, pl_headers)
        wb_pl.save(self.daily_pl_file)

    def _init_yearly_files(self):
        """Erstelle jährliche Excel-Dateien falls nicht vorhanden"""
        # Jahrestrades
        if not self.yearly_trades_file.exists():
            wb_trades = Workbook()
            ws_trades = wb_trades.active
            ws_trades.title = "Jahrestrades"

            trade_headers = ["Datum/Zeit", "Symbol", "Side", "Anzahl", "Preis", "Kommission", "Total Kosten"]
            self._write_headers(ws_trades, trade_headers)
            wb_trades.save(self.yearly_trades_file)

        # Jahres P/L
        if not self.yearly_pl_file.exists():
            wb_pl = Workbook()
            ws_pl = wb_pl.active
            ws_pl.title = "Jahres PL"

            pl_headers = ["Datum", "Symbol", "Total Aktien", "Total Trades", "Realisierter P/L", "Kommissionen", "Netto P/L"]
            self._write_headers(ws_pl, pl_headers)
            wb_pl.save(self.yearly_pl_file)

    def _write_headers(self, ws, headers: List[str]):
        """Schreibe Header-Zeile mit Formatierung"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Spaltenbreiten setzen
        column_widths = {
            "Datum/Zeit": 20,
            "Datum": 12,
            "Symbol": 10,
            "Side": 8,
            "Anzahl": 10,
            "Preis": 12,
            "Kommission": 12,
            "Total Kosten": 14,
            "Total Aktien": 12,
            "Total Trades": 12,
            "Realisierter P/L": 14,
            "Kommissionen": 12,
            "Netto P/L": 14
        }

        for col, header in enumerate(headers, 1):
            width = column_widths.get(header, 12)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header
        ws.freeze_panes = 'A2'

    def _format_number_swiss(self, value: float, decimals: int = 2) -> str:
        """Formatiert Zahl im Schweizer Format (1'000.00)"""
        if value is None:
            return "-"

        # Format mit Tausendertrennzeichen
        if decimals == 0:
            formatted = f"{value:,.0f}"
        else:
            formatted = f"{value:,.{decimals}f}"

        # Konvertiere zu Schweizer Format: , -> ' und . bleibt .
        return formatted.replace(",", "'")

    def add_trade(self, trade_data: dict):
        """
        Füge einen Trade zu allen relevanten Excel-Dateien hinzu

        Args:
            trade_data: Dictionary mit Trade-Informationen
                - symbol: Trading-Symbol
                - type: LONG oder SHORT
                - shares: Anzahl Aktien
                - entry_price: Einstiegspreis
                - price: Ausstiegspreis (Exit)
                - pnl: Netto P/L nach Kommission
                - commission: Total Kommission
                - scenario: Szenario-Name
                - level: Level-Nummer
        """
        timestamp = datetime.now()

        # Trade-Daten extrahieren
        symbol = trade_data.get('symbol', 'N/A')
        trade_type = trade_data.get('type', 'N/A')  # LONG oder SHORT
        shares = trade_data.get('shares', 0)
        entry_price = trade_data.get('entry_price', 0.0)
        exit_price = trade_data.get('price', 0.0)
        pnl = trade_data.get('pnl', 0.0)
        commission = trade_data.get('commission', 0.0)

        # Side bestimmen (Buy/Sell für Exit)
        # Bei LONG Exit = Sell, bei SHORT Exit = Buy
        if trade_type == 'LONG':
            side = 'Sell'
        else:
            side = 'Buy'

        # Total Kosten = Anzahl * Preis
        total_cost = shares * exit_price

        # Formatierte Werte
        datetime_str = timestamp.strftime("%Y-%m-%d %H-%M-%S")

        # 1. Zu Tagestrades hinzufügen
        self._add_to_daily_trades(datetime_str, symbol, side, shares, exit_price, commission, total_cost)

        # 2. Zu Jahrestrades hinzufügen
        self._add_to_yearly_trades(datetime_str, symbol, side, shares, exit_price, commission, total_cost)

        # 3. P/L-Statistiken pro Symbol aktualisieren
        self._update_symbol_stats(symbol, shares, pnl, commission)

    def _add_to_daily_trades(self, datetime_str: str, symbol: str, side: str,
                             shares: int, price: float, commission: float, total_cost: float):
        """Füge Trade zu Tagestrades hinzu"""
        try:
            wb = load_workbook(self.daily_trades_file)
            ws = wb.active

            # Finde nächste freie Zeile (vor Totals)
            next_row = ws.max_row + 1

            # Daten schreiben
            ws.cell(row=next_row, column=1, value=datetime_str)
            ws.cell(row=next_row, column=2, value=symbol)
            ws.cell(row=next_row, column=3, value=side)
            ws.cell(row=next_row, column=4, value=self._format_number_swiss(shares, 0))
            ws.cell(row=next_row, column=5, value=self._format_number_swiss(price))
            ws.cell(row=next_row, column=6, value=self._format_number_swiss(commission))
            ws.cell(row=next_row, column=7, value=self._format_number_swiss(total_cost))

            # Borders anwenden
            for col in range(1, 8):
                ws.cell(row=next_row, column=col).border = self.border

            wb.save(self.daily_trades_file)

        except Exception as e:
            print(f"Fehler beim Schreiben in Tagestrades: {e}")

    def _add_to_yearly_trades(self, datetime_str: str, symbol: str, side: str,
                              shares: int, price: float, commission: float, total_cost: float):
        """Füge Trade zu Jahrestrades hinzu"""
        try:
            wb = load_workbook(self.yearly_trades_file)
            ws = wb.active

            # Finde nächste freie Zeile
            next_row = ws.max_row + 1

            # Daten schreiben
            ws.cell(row=next_row, column=1, value=datetime_str)
            ws.cell(row=next_row, column=2, value=symbol)
            ws.cell(row=next_row, column=3, value=side)
            ws.cell(row=next_row, column=4, value=self._format_number_swiss(shares, 0))
            ws.cell(row=next_row, column=5, value=self._format_number_swiss(price))
            ws.cell(row=next_row, column=6, value=self._format_number_swiss(commission))
            ws.cell(row=next_row, column=7, value=self._format_number_swiss(total_cost))

            # Borders anwenden
            for col in range(1, 8):
                ws.cell(row=next_row, column=col).border = self.border

            wb.save(self.yearly_trades_file)

        except Exception as e:
            print(f"Fehler beim Schreiben in Jahrestrades: {e}")

    def _update_symbol_stats(self, symbol: str, shares: int, pnl: float, commission: float):
        """Aktualisiere P/L-Statistiken pro Symbol"""
        if symbol not in self.daily_symbol_stats:
            self.daily_symbol_stats[symbol] = {
                'total_shares': 0,
                'total_trades': 0,
                'realized_pnl': 0.0,
                'total_commissions': 0.0
            }

        stats = self.daily_symbol_stats[symbol]
        stats['total_shares'] += shares
        stats['total_trades'] += 1
        stats['realized_pnl'] += pnl + commission  # Brutto P/L
        stats['total_commissions'] += commission

    def write_daily_totals(self):
        """Schreibe Totals-Zeile in tägliche Trade-Datei"""
        try:
            wb = load_workbook(self.daily_trades_file)
            ws = wb.active

            if ws.max_row < 2:
                wb.save(self.daily_trades_file)
                return

            # Summen berechnen
            total_shares = 0
            total_commission = 0.0
            total_cost = 0.0

            for row in range(2, ws.max_row + 1):
                # Anzahl (Spalte 4)
                shares_val = ws.cell(row=row, column=4).value
                if shares_val and shares_val != "Total":
                    shares_str = str(shares_val).replace("'", "")
                    try:
                        total_shares += int(float(shares_str))
                    except:
                        pass

                # Kommission (Spalte 6)
                comm_val = ws.cell(row=row, column=6).value
                if comm_val:
                    comm_str = str(comm_val).replace("'", "")
                    try:
                        total_commission += float(comm_str)
                    except:
                        pass

                # Total Kosten (Spalte 7)
                cost_val = ws.cell(row=row, column=7).value
                if cost_val:
                    cost_str = str(cost_val).replace("'", "")
                    try:
                        total_cost += float(cost_str)
                    except:
                        pass

            # Totals-Zeile schreiben
            totals_row = ws.max_row + 1

            ws.cell(row=totals_row, column=1, value="Total")
            ws.cell(row=totals_row, column=4, value=self._format_number_swiss(total_shares, 0))
            ws.cell(row=totals_row, column=6, value=self._format_number_swiss(total_commission))
            ws.cell(row=totals_row, column=7, value=self._format_number_swiss(total_cost))

            # Totals formatieren
            for col in range(1, 8):
                cell = ws.cell(row=totals_row, column=col)
                cell.font = self.totals_font
                cell.fill = self.totals_fill
                cell.border = self.border

            wb.save(self.daily_trades_file)

        except Exception as e:
            print(f"Fehler beim Schreiben der Totals: {e}")

    def write_yearly_totals(self):
        """Schreibe Totals-Zeile in jährliche Trade-Datei"""
        try:
            wb = load_workbook(self.yearly_trades_file)
            ws = wb.active

            if ws.max_row < 2:
                wb.save(self.yearly_trades_file)
                return

            # Entferne alte Totals-Zeile falls vorhanden
            for row in range(ws.max_row, 1, -1):
                if ws.cell(row=row, column=1).value == "Total":
                    ws.delete_rows(row)
                    break

            # Summen berechnen
            total_shares = 0
            total_commission = 0.0
            total_cost = 0.0

            for row in range(2, ws.max_row + 1):
                # Anzahl (Spalte 4)
                shares_val = ws.cell(row=row, column=4).value
                if shares_val and shares_val != "Total":
                    shares_str = str(shares_val).replace("'", "")
                    try:
                        total_shares += int(float(shares_str))
                    except:
                        pass

                # Kommission (Spalte 6)
                comm_val = ws.cell(row=row, column=6).value
                if comm_val:
                    comm_str = str(comm_val).replace("'", "")
                    try:
                        total_commission += float(comm_str)
                    except:
                        pass

                # Total Kosten (Spalte 7)
                cost_val = ws.cell(row=row, column=7).value
                if cost_val:
                    cost_str = str(cost_val).replace("'", "")
                    try:
                        total_cost += float(cost_str)
                    except:
                        pass

            # Totals-Zeile schreiben
            totals_row = ws.max_row + 1

            ws.cell(row=totals_row, column=1, value="Total")
            ws.cell(row=totals_row, column=4, value=self._format_number_swiss(total_shares, 0))
            ws.cell(row=totals_row, column=6, value=self._format_number_swiss(total_commission))
            ws.cell(row=totals_row, column=7, value=self._format_number_swiss(total_cost))

            # Totals formatieren
            for col in range(1, 8):
                cell = ws.cell(row=totals_row, column=col)
                cell.font = self.totals_font
                cell.fill = self.totals_fill
                cell.border = self.border

            wb.save(self.yearly_trades_file)

        except Exception as e:
            print(f"Fehler beim Schreiben der jährlichen Totals: {e}")

    def write_daily_pl_summary(self):
        """Schreibe P/L-Zusammenfassung pro Symbol in Tages P/L"""
        try:
            wb = load_workbook(self.daily_pl_file)
            ws = wb.active

            # Lösche alle Daten ausser Header
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            if not self.daily_symbol_stats:
                wb.save(self.daily_pl_file)
                return

            date_str = self.session_start.strftime("%Y-%m-%d")
            row = 2

            # Zeile pro Symbol
            for symbol, stats in sorted(self.daily_symbol_stats.items()):
                netto_pl = stats['realized_pnl'] - stats['total_commissions']

                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=symbol)
                ws.cell(row=row, column=3, value=self._format_number_swiss(stats['total_shares'], 0))
                ws.cell(row=row, column=4, value=stats['total_trades'])
                ws.cell(row=row, column=5, value=self._format_number_swiss(stats['realized_pnl']))
                ws.cell(row=row, column=6, value=self._format_number_swiss(stats['total_commissions']))
                ws.cell(row=row, column=7, value=self._format_number_swiss(netto_pl))

                # Farbe für P/L
                pl_cell = ws.cell(row=row, column=7)
                if netto_pl >= 0:
                    pl_cell.font = self.positive_font
                else:
                    pl_cell.font = self.negative_font

                # Borders
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            wb.save(self.daily_pl_file)

        except Exception as e:
            print(f"Fehler beim Schreiben der Tages P/L: {e}")

    def write_yearly_pl_entry(self):
        """Füge Tages-P/L-Einträge zu Jahres P/L hinzu"""
        try:
            wb = load_workbook(self.yearly_pl_file)
            ws = wb.active

            if not self.daily_symbol_stats:
                wb.save(self.yearly_pl_file)
                return

            # Entferne alte Totals-Zeile falls vorhanden
            for row in range(ws.max_row, 1, -1):
                if ws.cell(row=row, column=1).value == "Total":
                    ws.delete_rows(row)
                    break

            date_str = self.session_start.strftime("%Y-%m-%d")
            next_row = ws.max_row + 1

            # Zeile pro Symbol für diesen Tag
            for symbol, stats in sorted(self.daily_symbol_stats.items()):
                netto_pl = stats['realized_pnl'] - stats['total_commissions']

                ws.cell(row=next_row, column=1, value=date_str)
                ws.cell(row=next_row, column=2, value=symbol)
                ws.cell(row=next_row, column=3, value=self._format_number_swiss(stats['total_shares'], 0))
                ws.cell(row=next_row, column=4, value=stats['total_trades'])
                ws.cell(row=next_row, column=5, value=self._format_number_swiss(stats['realized_pnl']))
                ws.cell(row=next_row, column=6, value=self._format_number_swiss(stats['total_commissions']))
                ws.cell(row=next_row, column=7, value=self._format_number_swiss(netto_pl))

                # Farbe für P/L
                pl_cell = ws.cell(row=next_row, column=7)
                if netto_pl >= 0:
                    pl_cell.font = self.positive_font
                else:
                    pl_cell.font = self.negative_font

                # Borders
                for col in range(1, 8):
                    ws.cell(row=next_row, column=col).border = self.border

                next_row += 1

            # Totals-Zeile berechnen und schreiben
            total_shares = 0
            total_trades = 0
            total_realized_pl = 0.0
            total_commissions = 0.0

            for row in range(2, ws.max_row + 1):
                # Total Aktien (Spalte 3)
                shares_val = ws.cell(row=row, column=3).value
                if shares_val:
                    shares_str = str(shares_val).replace("'", "")
                    try:
                        total_shares += int(float(shares_str))
                    except:
                        pass

                # Total Trades (Spalte 4)
                trades_val = ws.cell(row=row, column=4).value
                if trades_val:
                    try:
                        total_trades += int(trades_val)
                    except:
                        pass

                # Realisierter P/L (Spalte 5)
                pl_val = ws.cell(row=row, column=5).value
                if pl_val:
                    pl_str = str(pl_val).replace("'", "")
                    try:
                        total_realized_pl += float(pl_str)
                    except:
                        pass

                # Kommissionen (Spalte 6)
                comm_val = ws.cell(row=row, column=6).value
                if comm_val:
                    comm_str = str(comm_val).replace("'", "")
                    try:
                        total_commissions += float(comm_str)
                    except:
                        pass

            total_netto_pl = total_realized_pl - total_commissions

            # Totals-Zeile
            totals_row = ws.max_row + 1

            ws.cell(row=totals_row, column=1, value="Total")
            ws.cell(row=totals_row, column=3, value=self._format_number_swiss(total_shares, 0))
            ws.cell(row=totals_row, column=4, value=total_trades)
            ws.cell(row=totals_row, column=5, value=self._format_number_swiss(total_realized_pl))
            ws.cell(row=totals_row, column=6, value=self._format_number_swiss(total_commissions))
            ws.cell(row=totals_row, column=7, value=self._format_number_swiss(total_netto_pl))

            # Totals formatieren
            for col in range(1, 8):
                cell = ws.cell(row=totals_row, column=col)
                cell.font = self.totals_font
                cell.fill = self.totals_fill
                cell.border = self.border

            # Farbe für Netto P/L Total
            pl_total_cell = ws.cell(row=totals_row, column=7)
            if total_netto_pl >= 0:
                pl_total_cell.font = Font(bold=True, color="006100")
            else:
                pl_total_cell.font = Font(bold=True, color="9C0006")

            wb.save(self.yearly_pl_file)

        except Exception as e:
            print(f"Fehler beim Schreiben der Jahres P/L: {e}")

    def finalize_session(self):
        """
        Finalisiere die Session - schreibe alle Totals und P/L-Zusammenfassungen
        Wird aufgerufen wenn die App geschlossen wird
        """
        # Tägliche Trade-Totals
        self.write_daily_totals()

        # Jährliche Trade-Totals
        self.write_yearly_totals()

        # Tägliche P/L-Zusammenfassung
        self.write_daily_pl_summary()

        # Jährliche P/L-Einträge
        self.write_yearly_pl_entry()

    def save_intermediate(self):
        """
        Speichere aktuelle P/L-Daten ohne Session zu beenden.
        Kann sicher mehrfach aufgerufen werden (für Auto-Save und manuelles Speichern).
        Schützt vor Datenverlust bei Absturz.
        """
        # Tägliche P/L-Zusammenfassung (löscht und schreibt neu - sicher für mehrfaches Aufrufen)
        self.write_daily_pl_summary()

        # Temporäre Totals für Tages-Trades (überschreibt vorherige)
        self._write_intermediate_daily_totals()

    def _write_intermediate_daily_totals(self):
        """Schreibe temporäre Totals-Zeile in Tages-Trades (wird bei jedem Save überschrieben)"""
        try:
            wb = load_workbook(self.daily_trades_file)
            ws = wb.active

            # Entferne vorherige Totals-Zeile falls vorhanden
            for row in range(ws.max_row, 1, -1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and ("TOTAL" in str(cell_value).upper() or "ZWISCHENSUMME" in str(cell_value).upper()):
                    ws.delete_rows(row)
                    break

            # Berechne Totals
            if ws.max_row < 2:
                wb.save(self.daily_trades_file)
                return

            total_shares = 0
            total_cost = 0.0
            total_commission = 0.0

            for row in range(2, ws.max_row + 1):
                try:
                    shares_val = ws.cell(row=row, column=4).value
                    cost_val = ws.cell(row=row, column=7).value
                    comm_val = ws.cell(row=row, column=6).value

                    if shares_val:
                        shares_str = str(shares_val).replace("'", "").replace(",", "")
                        total_shares += int(float(shares_str)) if shares_str else 0
                    if cost_val:
                        cost_str = str(cost_val).replace("'", "").replace(",", "")
                        total_cost += float(cost_str) if cost_str else 0
                    if comm_val:
                        comm_str = str(comm_val).replace("'", "").replace(",", "")
                        total_commission += float(comm_str) if comm_str else 0
                except:
                    pass

            # Schreibe Zwischensumme
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1, value="ZWISCHENSUMME")
            ws.cell(row=total_row, column=4, value=self._format_number_swiss(total_shares, 0))
            ws.cell(row=total_row, column=6, value=self._format_number_swiss(total_commission))
            ws.cell(row=total_row, column=7, value=self._format_number_swiss(total_cost))

            # Formatierung
            for col in range(1, 8):
                cell = ws.cell(row=total_row, column=col)
                cell.font = self.totals_font
                cell.fill = self.totals_fill
                cell.border = self.border

            wb.save(self.daily_trades_file)

        except Exception as e:
            print(f"Fehler beim Schreiben der Zwischensumme: {e}")

        print(f"Trading Logs finalisiert:")
        print(f"  - Tagestrades: {self.daily_trades_file}")
        print(f"  - Jahrestrades: {self.yearly_trades_file}")
        print(f"  - Tages P/L: {self.daily_pl_file}")
        print(f"  - Jahres P/L: {self.yearly_pl_file}")
