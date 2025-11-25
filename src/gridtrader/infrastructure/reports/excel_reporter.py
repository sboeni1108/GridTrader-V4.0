"""
Excel Reporter für GridTrader V2.0
Erstellt deutsche Reports im Schweizer Format
"""
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import locale

from gridtrader.domain.services.backtest_engine import BacktestResult
from gridtrader.domain.models.order import Trade


class ExcelReporter:
    """
    Erstellt Excel-Reports in deutschem Format (CH)
    """
    
    def __init__(self, output_dir: str = "reports/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Versuche Schweizer Locale zu setzen
        try:
            locale.setlocale(locale.LC_ALL, 'de_CH.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'German_Switzerland.1252')
            except:
                pass  # Fallback auf System-Default
    
    def create_backtest_report(
        self, 
        result: BacktestResult,
        filename: Optional[str] = None
    ) -> Path:
        """
        Erstellt Backtest-Report
        
        Args:
            result: Backtest-Ergebnis
            filename: Optionaler Dateiname
            
        Returns:
            Pfad zur erstellten Datei
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Backtest_{result.symbol}_{result.side}_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Workbook erstellen
        wb = Workbook()
        
        # Sheets erstellen
        self._create_overview_sheet(wb, result)
        self._create_trades_sheet(wb, result)
        self._create_statistics_sheet(wb, result)
        self._create_levels_sheet(wb, result)
        
        # Standard-Sheet entfernen
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Formatierung anwenden
        self._apply_formatting(wb)

        # Speichern
        wb.save(filepath)

        print(f"📊 Excel-Report erstellt: {filepath}")
        return filepath
    
    def _create_overview_sheet(self, wb: Workbook, result: BacktestResult):
        """Erstellt Übersichts-Sheet nach Original-Format"""
        ws = wb.create_sheet("Übersicht")

        # Styles definieren
        section_header_font = Font(bold=True)
        section_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Header (Zeile 1)
        ws["A1"] = "Parameter"
        ws["A1"].font = Font(bold=True, size=14)

        # Titel (Zeile 2)
        ws["A2"] = f"GRID TRADING BACKTEST - {result.symbol} {result.side.upper()}"

        # Leerzeile 3
        row = 4

        # 📊 ANALYSE-UMFANG
        ws[f"A{row}"] = "📊 ANALYSE-UMFANG"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        row += 1

        analyse_data = [
            ("Symbol", result.symbol),
            ("Seite", result.side.upper()),
            ("Analysezeitraum", f"{result.trading_days} Tage"),
            ("Start", result.start_date.strftime("%d.%m.%Y")),
            ("Ende", result.end_date.strftime("%d.%m.%Y")),
        ]

        for label, value in analyse_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1  # Leerzeile

        # 💰 PERFORMANCE
        ws[f"A{row}"] = "💰 PERFORMANCE"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        row += 1

        performance_data = [
            ("Startkapital", self._format_currency(result.starting_capital)),
            ("Endkapital", self._format_currency(result.ending_capital)),
            ("Netto-Gewinn", self._format_currency(result.total_return)),
            ("ROI", self._format_percent(result.total_return_pct)),
            ("Sharpe Ratio", f"{result.sharpe_ratio:.2f}"),
            ("Max Drawdown", self._format_percent(result.max_drawdown_pct)),
        ]

        for label, value in performance_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1  # Leerzeile

        # 📈 TRADING-STATISTIK
        ws[f"A{row}"] = "📈 TRADING-STATISTIK"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        row += 1

        trading_data = [
            ("Erfolgreiche Trades", str(result.winning_trades)),
            ("Verlorene Trades", str(result.losing_trades)),
            ("Total Trades", str(result.total_trades)),
            ("Win Rate", self._format_percent(Decimal(str(result.win_rate)))),
            ("Ø Gewinn", self._format_currency(result.avg_win)),
            ("Ø Verlust", self._format_currency(result.avg_loss)),
            ("Profit Faktor", f"{result.profit_factor:.2f}"),
        ]

        for label, value in trading_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1  # Leerzeile

        # 🏆 GRID-DETAILS
        ws[f"A{row}"] = "🏆 GRID-DETAILS"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        row += 1

        grid_data = [
            ("Levels ausgelöst", str(result.levels_triggered)),
            ("Levels abgeschlossen", str(result.levels_completed)),
            ("Ø Zeit im Trade (Std)", f"{result.avg_time_in_trade:.1f}"),
        ]

        for label, value in grid_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
    
    def _create_trades_sheet(self, wb: Workbook, result: BacktestResult):
        """Erstellt Trades-Sheet"""
        ws = wb.create_sheet("Trades")
        
        # Headers
        headers = [
            "Trade Nr.", "Symbol", "Seite", "Menge",
            "Einstiegspreis", "Ausstiegspreis", "Gebühren",
            "Brutto P&L", "Netto P&L", "Datum/Zeit"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Trade-Daten
        row = 2
        for i, trade in enumerate(result.trades, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=trade.symbol)
            ws.cell(row=row, column=3, value=trade.side.value)
            ws.cell(row=row, column=4, value=trade.quantity)
            ws.cell(row=row, column=5, value=self._format_price(trade.price))
            ws.cell(row=row, column=6, value="-")  # Exit price in nächster Zeile
            ws.cell(row=row, column=7, value=self._format_currency(trade.commission))
            
            if trade.realized_pnl:
                ws.cell(row=row, column=8, value=self._format_currency(trade.realized_pnl + trade.commission))
                ws.cell(row=row, column=9, value=self._format_currency(trade.realized_pnl))
            
            ws.cell(row=row, column=10, value=trade.executed_at.strftime("%d.%m.%Y %H:%M:%S"))
            row += 1
        
        # Auto-Filter
        ws.auto_filter.ref = f"A1:J{row-1}"
        
        # Spaltenbreiten
        for col in range(1, 11):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def _create_statistics_sheet(self, wb: Workbook, result: BacktestResult):
        """Erstellt Statistik-Sheet"""
        ws = wb.create_sheet("Statistiken")
        
        # Detaillierte Statistiken
        ws["A1"] = "DETAILLIERTE STATISTIKEN"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Grid-Statistiken
        ws["A3"] = "Grid-Statistiken"
        ws["A3"].font = Font(bold=True, size=12)
        
        grid_stats = [
            ("Levels konfiguriert", "-"),  # TODO: aus Template
            ("Levels ausgelöst", str(result.levels_triggered)),
            ("Levels abgeschlossen", str(result.levels_completed)),
            ("Ø Zeit im Trade (Std)", f"{result.avg_time_in_trade:.1f}"),
        ]
        
        row = 4
        for label, value in grid_stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        
        # Zeitraum
        ws["D3"] = "Zeitraum"
        ws["D3"].font = Font(bold=True, size=12)
        
        period_stats = [
            ("Start", result.start_date.strftime("%d.%m.%Y")),
            ("Ende", result.end_date.strftime("%d.%m.%Y")),
            ("Handelstage", str(result.trading_days)),
        ]
        
        row = 4
        for label, value in period_stats:
            ws[f"D{row}"] = label
            ws[f"E{row}"] = value
            row += 1
    
    def _create_levels_sheet(self, wb: Workbook, result: BacktestResult):
        """Erstellt Levels-Sheet (für nicht ausgeführte Stufen)"""
        ws = wb.create_sheet("Nicht ausgeführte Stufen")
        
        # Headers
        headers = ["Symbol", "Seite", "Level", "Geplanter Preis", "Menge", "Grund"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Platzhalter für nicht ausgeführte Levels
        ws.cell(row=2, column=1, value="(Keine Daten verfügbar)")
        
    def _format_currency(self, value: Decimal) -> str:
        """Formatiert als Währung (CH-Format)"""
        if value is None:
            return "-"
        # Schweizer Format: 1'234.56
        return f"CHF {value:,.2f}".replace(",", "'").replace(".", ",")
    
    def _format_price(self, value: Decimal) -> str:
        """Formatiert als Preis"""
        if value is None:
            return "-"
        return f"{value:.2f}".replace(".", ",")
    
    def _format_percent(self, value: Decimal) -> str:
        """Formatiert als Prozent"""
        if value is None:
            return "-"
        return f"{value:.2f}%".replace(".", ",")

    def _apply_formatting(self, wb: Workbook):
        """Wendet professionelle Formatierung auf alle Sheets an"""

        # Definiere Styles
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Positive/Negative Farben
        positive_font = Font(color="006100")
        negative_font = Font(color="9C0006")

        # Formatiere jedes Sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # === SPEZIELLE FORMATIERUNG PRO SHEET ===

            if sheet_name == 'Trades':
                # Freeze erste Zeile
                sheet.freeze_panes = 'A2'

                # Conditional Formatting für P&L Spalten (Spalte 8 und 9)
                for row_idx in range(2, sheet.max_row + 1):
                    for col_idx in [8, 9]:  # Brutto P&L und Netto P&L
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            value_str = str(cell.value)
                            if '-' in value_str:
                                cell.font = negative_font
                            elif 'CHF' in value_str or any(c.isdigit() for c in value_str):
                                cell.font = positive_font

            elif sheet_name == 'Nicht ausgeführte Stufen':
                # Freeze erste Zeile
                sheet.freeze_panes = 'A2'

            # === ALLGEMEINE FORMATIERUNG ===

            # Auto-Spaltenbreite
            for column in sheet.columns:
                max_length = 0

                # Hole column_letter sicher
                first_cell = column[0]

                if hasattr(first_cell, 'column_letter'):
                    column_letter = first_cell.column_letter
                elif hasattr(first_cell, 'column'):
                    column_letter = get_column_letter(first_cell.column)
                else:
                    continue

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                            # Rahmen hinzufügen für Datenzellen
                            if cell.row > 1:
                                cell.border = border
                    except:
                        pass

                # Setze Spaltenbreite (min 10, max 50)
                if max_length > 0:
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    try:
                        sheet.column_dimensions[column_letter].width = adjusted_width
                    except:
                        pass

            # Zeilenhöhe für Header
            if sheet.max_row > 0:
                sheet.row_dimensions[1].height = 25


def create_comparison_report(
    long_result: BacktestResult,
    short_result: BacktestResult,
    output_dir: str = "reports/output"
) -> Path:
    """
    Erstellt Vergleichsreport Long vs Short
    """
    reporter = ExcelReporter(output_dir)
    
    # TODO: Implementierung für Vergleichsreport
    
    pass
